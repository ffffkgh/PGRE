import torch
import numpy as np
import matplotlib.pyplot as plt
from collections import Counter
import os
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def append_eval_excel(xlsx_path, dataset, R, T, N, results: dict, threshold=0.5, sheet_name="Eval"):
    """
    将一次评估结果追加写入 Excel（一次一行）。
    兼容 pandas 新版本：不再设置 writer.book。
    """
    # 取数 & 行构造
    f1_key = f"F1@{threshold}"
    rec_key = f"Recall@{threshold}"

    def _get(d, k):
        v = d.get(k, np.nan)
        try:
            return float(v)
        except Exception:
            return np.nan

    row = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "dataset": dataset,
        "R": R, "T": T, "N": N,
        "ROC-AUC": _get(results, "ROC-AUC"),
        "PR-AUC": _get(results, "PR-AUC"),
        f"F1@{threshold}": _get(results, f1_key),
        f"Recall@{threshold}": _get(results, rec_key),
        "Best-F1": _get(results, "Best-F1"),
        "Best-Precision": _get(results, "Best-Precision"),
        "Best-Recall": _get(results, "Best-Recall"),
        "Best-threshold": _get(results, "Best-threshold"),
    }
    df = pd.DataFrame([row])

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    if not os.path.exists(xlsx_path):
        # 首次创建文件
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # 文件已存在：判断工作表是否存在，并决定写入起始行
    book = load_workbook(xlsx_path)
    if sheet_name in book.sheetnames:
        startrow = book[sheet_name].max_row  # 现有最后一行
        # 追加（不写表头）
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    else:
        # 新建工作表（写表头）
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
def _infer_TN(paths):
    """从多个文件中推断 N、T（实体与时间的最大索引+1）。"""
    h_max = t_max = time_max = -1
    for p in paths:
        with open(p, 'r', encoding='utf-8') as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                parts = s.split()
                if len(parts) < 4:
                    continue
                try:
                    h = int(parts[0]); r = int(parts[1]); tail = int(parts[2]); tm = int(parts[3])
                except ValueError:
                    continue
                if h > h_max: h_max = h
                if tail > t_max: t_max = tail
                if tm > time_max: time_max = tm
    N = max(h_max, t_max) + 1 if h_max >= 0 or t_max >= 0 else 0
    T = time_max + 1 if time_max >= 0 else 0
    return T, N

def load_4d_array_from_txts(train_path, val_path, test_path, R):
    """
    - 仅用 TRAIN 统计关系频次，选出 Top-R 并建立 rel_old2new（旧ID->新ID，0..R_sel-1）
    - 从 train+val+test 推断 T 与 N
    - 用同一映射将 train+val+test 的四元组统一填入 (R, T, N, N) 的张量

    返回:
      adjacency_4d: np.ndarray, shape (R, T, N, N)
      rel_old2new: dict, 旧关系ID -> 新关系ID（仅含 Top-R）
      T: int, 推断的时间片数
      N: int, 推断的实体数
    """
    # 1) 仅在 train 统计关系频次，选 Top-R
    rel_cnt = Counter()
    with open(train_path, 'r', encoding='utf-8') as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            parts = s.split()
            if len(parts) < 5:
                continue
            try:
                r = int(parts[1])
            except ValueError:
                continue
            rel_cnt[r] += 1

    # 频次降序、旧ID升序打破并列
    top_relations = [rid for rid, _ in sorted(rel_cnt.items(), key=lambda kv: (-kv[1], kv[0]))[:R]]
    R_sel = min(R, len(top_relations))
    rel_old2new = {old_r: new_r for new_r, old_r in enumerate(top_relations[:R_sel])}

    # 2) 从三个文件推断 T、N
    T, N = _infer_TN([train_path, val_path, test_path])
    if T == 0 or N == 0:
        # 防御：如果文件为空或解析失败，给出空张量
        return np.zeros((R, 0, 0, 0), dtype=int), rel_old2new, T, N

    # 3) 初始化张量
    adjacency_4d = np.zeros((R, T, N, N), dtype=int)

    # 4) 用同一映射填充 train + val + test
    for file_path in [train_path, val_path, test_path]:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                parts = s.split()
                if len(parts) < 5:
                    continue
                try:
                    h = int(parts[0]); r = int(parts[1]); x = int(parts[2]); t = int(parts[3])
                except ValueError:
                    continue

                # 仅保留 Top-R 关系
                if r not in rel_old2new:
                    continue
                r_new = rel_old2new[r]

                # 边界检查（按推断出的 T、N）
                if 0 <= r_new < R and 0 <= t < T and 0 <= h < N and 0 <= x < N:
                    adjacency_4d[r_new, t, h, x] = 1

    return adjacency_4d, rel_old2new, T, N

def visualize_relation_matrix(train_file, relation_id, time_id, node_count=250):
    """
    从 train_file 中读取三元组 (head, r, tail, time, 0)，
    构造一个 node_count x node_count 的矩阵:
        matrix[head, tail] = 1  (若 r == relation_id 且 time == time_id)
    最后用 Matplotlib 可视化。
    """
    # 1. 初始化 NxN 矩阵，全 0
    matrix = np.zeros((node_count, node_count), dtype=int)

    # 2. 读取文件
    with open(train_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) < 5:
                continue
            h_str, r_str, t_str, time_str, _ = parts
            h = int(h_str)
            r = int(r_str)
            t = int(t_str)
            time_ = int(time_str)

            # 如果满足指定关系 & 时间片，则设置 matrix[h, t] = 1
            if r == relation_id and time_ == time_id:
                # 这里假设 h, t 已经是 0~249 的 ID
                # 如果不是，需要做映射或检查
                if 0 <= h < node_count and 0 <= t < node_count:
                    matrix[h, t] = 1

    # 3. 用 Matplotlib 可视化
    plt.figure(figsize=(6,6))
    plt.imshow(matrix, cmap='Greys', interpolation='none', origin='upper')
    plt.title(f"Relation={relation_id}, Time={time_id}")
    plt.xlabel("Tail Node")
    plt.ylabel("Head Node")
    plt.colorbar(label="1=exist, 0=none")
    plt.show()

def inspect_data_test(data_test, remove_self_loop=True, top_edges=10):
    """
    data_test: torch.Tensor 或 np.ndarray, 形状 [R, T, N, N]
    remove_self_loop: 是否在统计密度时忽略自环
    top_edges: 打印前多少条非零边示例
    """
    # --- to numpy ---
    if isinstance(data_test, torch.Tensor):
        arr = data_test.detach().cpu().numpy()
        dtype = str(data_test.dtype)
        device = str(data_test.device)
    else:
        arr = np.asarray(data_test)
        dtype = str(arr.dtype)
        device = "numpy"

    if arr.ndim != 4:
        print(f"❌ 维度错误: 期望 4 维 [R,T,N,N]，实际 {arr.shape}")
        return

    R, T, N, N2 = arr.shape
    print(f"[info] shape={arr.shape} (R={R}, T={T}, N={N}), dtype={dtype}, device={device}")
    assert N == N2, "最后两维应该相等（N×N）"

    # --- 基本统计 ---
    vmin, vmax = arr.min(initial=0), arr.max(initial=0)
    uniq_small = np.unique(arr[:,:,:,:].ravel()[:100000])  # 采样看下唯一值（避免超大内存）
    print(f"[values] min={vmin}, max={vmax}, sample_unique={uniq_small[:10]}{' ...' if uniq_small.size>10 else ''}")

    # 判断是否二值
    is_binary = np.all(np.isin(uniq_small, [0,1])) and (vmax <= 1)
    print(f"[values] binary_like={is_binary}")

    # --- 正例统计（计数>0 即为正）---
    pos_mask = (arr > 0)
    total_pos = int(pos_mask.sum())
    total_cells = R * T * (N * (N-1) if remove_self_loop else N*N)
    if remove_self_loop:
        # 去掉对角线后再算稀疏度
        mask = np.ones((N, N), dtype=bool)
        np.fill_diagonal(mask, False)
        pos_no_diag = int((pos_mask * mask).sum())
        density = pos_no_diag / total_cells if total_cells > 0 else 0.0
        print(f"[global] positives(off-diag)={pos_no_diag}, density(off-diag)={density:.6e}")
    else:
        density = total_pos / (R*T*N*N)
        print(f"[global] positives={total_pos}, density={density:.6e}")

    # --- 按关系/时间统计 ---
    pos_by_r = pos_mask.sum(axis=(1,2,3))           # [R]
    pos_by_t = pos_mask.sum(axis=(0,2,3))           # [T]
    print("[by-r] 非零边数（前10个）:", pos_by_r[:10])
    print("[by-t] 非零边数（全部）  :", pos_by_t)

    # --- (r,t) 级别诊断：全 0 / 仅自环 / 有非自环 ---
    degenerate_zero = []    # 完全没有边
    diag_only = []          # 只有自环
    valid_blocks = 0

    eye = np.eye(N, dtype=bool)
    for r in range(R):
        for t in range(T):
            A = arr[r, t]
            pos = int((A > 0).sum())
            if pos == 0:
                degenerate_zero.append((r, t))
                continue
            diag = int(A[eye].sum())
            off  = pos - diag
            if off == 0:
                diag_only.append((r, t))
            else:
                valid_blocks += 1

    print(f"[blocks] 总块数={R*T}, 有非自环块={valid_blocks}, 全0块={len(degenerate_zero)}, 仅自环块={len(diag_only)}")
    if len(degenerate_zero) > 0:
        print("  例子(全0块)前5个:", degenerate_zero[:5])
    if len(diag_only) > 0:
        print("  例子(仅自环)前5个:", diag_only[:5])

    # --- 抽样打印非零边示例 ---
    # 注意：当数据很大时，np.argwhere 可能较慢；这里只取前 top_edges 条
    if total_pos > 0:
        idx = np.argwhere(arr > 0)
        print(f"[samples] 非零边示例（最多 {top_edges} 条，格式 r,t,h,j,val）:")
        for k in range(min(top_edges, idx.shape[0])):
            r, t, h, j = idx[k]
            print(f"  ({r},{t},{h},{j}) -> {arr[r,t,h,j]}")

    # --- 友好提示 ---
    if total_pos == 0:
        print("❗ data_test 全 0：请检查 (1) R/T/N 是否太小导致越界丢样本；(2) 测试时间片是否真没有边；(3) 是否在评估前把自环去掉导致全 0。")
    elif not is_binary:
        print("ℹ️ 检测到计数标签（>1），若做二分类评估请先二值化：labels = (data_test > 0).astype(int)")

def main():
    train_file = "data/WIKI/train_sliced.txt"
    relation_id = 1
    time_id = 6
    node_count = 250

    visualize_relation_matrix(train_file, relation_id, time_id, node_count)

if __name__ == "__main__":
    main()
